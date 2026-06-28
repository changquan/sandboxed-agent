import base64
import json
import os
from pathlib import Path

import anthropic
import chainlit as cl

from src.sandbox import get_sandbox_session

_COMPLIANCE_DIR = "/compliance"
_REGISTRY_PATH = f"{_COMPLIANCE_DIR}/registry.json"
_SCRIPT_PATH = f"{_COMPLIANCE_DIR}/compliance_check.py"
_REPORT_PATH = f"{_COMPLIANCE_DIR}/compliance_report.xlsx"

_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
_PDF_EXTENSIONS = {".pdf"}
_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}

_EXCEL_PARSE_TEMPLATE = """
import pandas as pd, json, sys
df = pd.read_excel("{path}", engine="openpyxl")
df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

COL_MAP_TRADES = {{
    "ticker": ["ticker", "symbol", "stock", "security"],
    "isin": ["isin"],
    "direction": ["direction", "side", "type", "action", "buy_sell"],
    "quantity": ["quantity", "qty", "shares", "units", "volume"],
    "trade_date": ["trade_date", "date", "transaction_date", "settlement_date"],
    "account": ["account", "account_number", "acct"],
}}

COL_MAP_PC = {{
    "ticker": ["ticker", "symbol", "security"],
    "isin": ["isin"],
    "direction": ["direction", "side", "type"],
    "approved_quantity": ["approved_quantity", "approved_qty", "qty_approved", "quantity", "qty"],
    "approval_date": ["approval_date", "approved_date", "date_approved", "request_date"],
    "expiry_date": ["expiry_date", "expiry", "expires", "valid_until", "valid_to"],
    "employee": ["employee", "employee_name", "name", "person"],
}}

schema = COL_MAP_{role}

def find_col(df, aliases):
    for a in aliases:
        if a in df.columns:
            return a
    return None

records = []
for _, row in df.iterrows():
    rec = {{}}
    for field, aliases in schema.items():
        col = find_col(df, aliases)
        val = str(row[col]).strip() if col and col in df.columns else ""
        if val.lower() in ("nan", "none", ""):
            val = ""
        rec[field] = val
    if any(rec.values()):
        records.append(rec)

print(json.dumps(records))
"""

_PDF_PARSE_TEMPLATE = """
import pdfplumber, json, re

records = []
with pdfplumber.open("{path}") as pdf:
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            if not table:
                continue
            headers = [str(h).strip().lower().replace(" ", "_") for h in table[0]]
            for row in table[1:]:
                rec = dict(zip(headers, [str(c).strip() if c else "" for c in row]))
                records.append(rec)
        if not tables:
            text = page.extract_text() or ""
            for line in text.splitlines():
                parts = re.split(r"\\s{{2,}}|\\t", line.strip())
                if len(parts) >= 3:
                    records.append({{"raw": line.strip()}})

print(json.dumps(records))
"""

_IMAGE_PROMPT = """Extract all trades from this broker statement image.
Return a JSON array where each element has these exact keys:
  ticker, isin, direction (buy or sell), quantity (number), trade_date (YYYY-MM-DD), account
If a field is not present, use an empty string.
Return ONLY the JSON array, no explanation."""


def _sbx(session):
    """Get the underlying e2b AsyncSandbox. Adjust if your SDK version differs."""
    for attr in ("sandbox", "_sandbox", "e2b"):
        if hasattr(session, attr):
            candidate = getattr(session, attr)
            if hasattr(candidate, "files") or hasattr(candidate, "filesystem"):
                return candidate
    return session


async def _sbx_write(session, path: str, data: bytes) -> None:
    sbx = _sbx(session)
    if hasattr(sbx, "files"):
        await sbx.files.write(path, data)
    elif hasattr(sbx, "filesystem"):
        await sbx.filesystem.write(path, data)
    else:
        raise RuntimeError(f"Cannot write to sandbox: unknown session type {type(sbx)}")


async def _sbx_read(session, path: str) -> bytes:
    sbx = _sbx(session)
    if hasattr(sbx, "files"):
        return await sbx.files.read(path)
    elif hasattr(sbx, "filesystem"):
        return await sbx.filesystem.read(path)
    else:
        raise RuntimeError(f"Cannot read from sandbox: unknown session type {type(sbx)}")


async def _sbx_run(session, cmd: str) -> tuple[str, str, int]:
    """Returns (stdout, stderr, exit_code)."""
    sbx = _sbx(session)
    if hasattr(sbx, "commands"):
        result = await sbx.commands.run(cmd, timeout=120)
        return result.stdout, result.stderr, result.exit_code
    elif hasattr(sbx, "process"):
        proc = await sbx.process.start(cmd)
        out = await proc.wait()
        return out.stdout, out.stderr, out.exit_code
    else:
        raise RuntimeError(f"Cannot run command in sandbox: unknown session type {type(sbx)}")


async def _ensure_dir(session) -> None:
    await _sbx_run(session, f"mkdir -p {_COMPLIANCE_DIR}")


async def _read_registry(session) -> dict:
    try:
        data = await _sbx_read(session, _REGISTRY_PATH)
        return json.loads(data)
    except Exception:
        return {}


async def _write_registry(session, registry: dict) -> None:
    await _sbx_write(session, _REGISTRY_PATH, json.dumps(registry).encode())


async def _parse_excel_in_sandbox(session, sandbox_path: str, role: str) -> list:
    role_key = "TRADES" if role == "broker_statement" else "PC"
    script = _EXCEL_PARSE_TEMPLATE.format(path=sandbox_path, role=role_key)
    tmp_script = f"{_COMPLIANCE_DIR}/_parse_excel.py"
    await _sbx_write(session, tmp_script, script.encode())
    stdout, stderr, code = await _sbx_run(session, f"pip install openpyxl -q && python3 {tmp_script}")
    if code != 0:
        raise RuntimeError(f"Excel parse failed: {stderr}")
    return json.loads(stdout)


async def _parse_pdf_in_sandbox(session, sandbox_path: str) -> list:
    script = _PDF_PARSE_TEMPLATE.format(path=sandbox_path)
    tmp_script = f"{_COMPLIANCE_DIR}/_parse_pdf.py"
    await _sbx_write(session, tmp_script, script.encode())
    stdout, stderr, code = await _sbx_run(session, f"pip install pdfplumber -q && python3 {tmp_script}")
    if code != 0:
        raise RuntimeError(f"PDF parse failed: {stderr}")
    return json.loads(stdout)


async def _parse_image_with_claude(local_path: str) -> list:
    image_data = Path(local_path).read_bytes()
    b64 = base64.standard_b64encode(image_data).decode()
    ext = Path(local_path).suffix.lower().lstrip(".")
    media_type_map = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "gif": "image/gif", "webp": "image/webp",
    }
    media_type = media_type_map.get(ext, "image/png")

    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": _IMAGE_PROMPT},
            ],
        }],
    )
    text = response.content[0].text.strip()
    if text.startswith("```"):
        text = "\n".join(text.split("\n")[1:])
        text = text.rstrip("`").strip()
    return json.loads(text)


async def _generate_report(session, result: dict) -> bytes:
    result_json_path = f"{_COMPLIANCE_DIR}/_result.json"
    await _sbx_write(session, result_json_path, json.dumps(result).encode())
    report_script = f"""
import json, openpyxl
from openpyxl.styles import Font, PatternFill

with open("{result_json_path}") as f:
    result = json.load(f)

wb = openpyxl.Workbook()

def header_row(ws, cols):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

ws1 = wb.active
ws1.title = "Discrepancies"
header_row(ws1, ["Ticker", "Direction", "Traded Qty", "Approved Qty",
                  "Trade Date", "Approval Date", "Expiry Date", "Reason(s)"])
for item in result["discrepancies"]:
    t = item["trade"]
    pc = item["preclearance"] or {{}}
    ws1.append([
        t.get("ticker", ""), t.get("direction", ""), t.get("quantity", ""),
        pc.get("approved_quantity", "N/A"), t.get("trade_date", ""),
        pc.get("approval_date", "N/A"), pc.get("expiry_date", "N/A"),
        ", ".join(item["reasons"]),
    ])

ws2 = wb.create_sheet("Clean Trades")
header_row(ws2, ["Ticker", "Direction", "Traded Qty", "Approved Qty",
                  "Trade Date", "Approval Date", "Expiry Date"])
for item in result["clean_trades"]:
    t = item["trade"]
    pc = item["preclearance"]
    ws2.append([
        t.get("ticker", ""), t.get("direction", ""), t.get("quantity", ""),
        pc.get("approved_quantity", ""), t.get("trade_date", ""),
        pc.get("approval_date", ""), pc.get("expiry_date", ""),
    ])

ws3 = wb.create_sheet("Exempt Trades")
header_row(ws3, ["Ticker", "ISIN", "Direction", "Quantity", "Trade Date", "Exempt Reason"])
for t in result["exempt_trades"]:
    ws3.append([
        t.get("ticker", ""), t.get("isin", ""), t.get("direction", ""),
        t.get("quantity", ""), t.get("trade_date", ""), t.get("exempt_reason", ""),
    ])

wb.save("{_REPORT_PATH}")
print("ok")
"""
    tmp = f"{_COMPLIANCE_DIR}/_gen_report.py"
    await _sbx_write(session, tmp, report_script.encode())
    stdout, stderr, code = await _sbx_run(session, f"python3 {tmp}")
    if code != 0:
        raise RuntimeError(f"Report generation failed: {stderr}")
    return await _sbx_read(session, _REPORT_PATH)


async def register_compliance_file(filename: str, role: str) -> str:
    """
    Register an uploaded file for compliance checking.

    Args:
        filename: Name of the uploaded file (must match a file the user uploaded this turn).
        role: One of 'broker_statement', 'preclearance', or 'index_reference'.
    """
    valid_roles = {"broker_statement", "preclearance", "index_reference"}
    if role not in valid_roles:
        return f"Invalid role '{role}'. Must be one of: {', '.join(sorted(valid_roles))}"

    uploaded: dict = cl.user_session.get("uploaded_files", {})
    if filename not in uploaded:
        available = list(uploaded.keys()) or ["(none)"]
        return f"File '{filename}' not found in this session. Available: {', '.join(available)}"

    local_path = uploaded[filename]
    ext = Path(filename).suffix.lower()
    session = get_sandbox_session()
    if session is None:
        return "Sandbox not available. Please start a new chat."

    await _ensure_dir(session)

    sandbox_path = f"{_COMPLIANCE_DIR}/{role}{ext}"
    data = Path(local_path).read_bytes()
    await _sbx_write(session, sandbox_path, data)

    if ext in _EXCEL_EXTENSIONS:
        records = await _parse_excel_in_sandbox(session, sandbox_path, role)
    elif ext in _PDF_EXTENSIONS:
        records = await _parse_pdf_in_sandbox(session, sandbox_path)
    elif ext in _IMAGE_EXTENSIONS:
        records = await _parse_image_with_claude(local_path)
    elif ext == ".json" and role == "index_reference":
        records = json.loads(data)
    else:
        return f"Unsupported file type '{ext}' for role '{role}'."

    json_path = f"{_COMPLIANCE_DIR}/{role}.json"
    await _sbx_write(session, json_path, json.dumps(records).encode())

    registry = await _read_registry(session)
    registry[role] = {"filename": filename, "records": len(records), "json_path": json_path}
    await _write_registry(session, registry)

    return (
        f"Registered '{filename}' as {role.replace('_', ' ')}. "
        f"Extracted {len(records)} record(s). "
        f"Registry now has: {', '.join(registry.keys())}."
    )


async def run_compliance_check() -> str:
    """
    Run the compliance check comparing registered broker statement against pre-clearance declarations.
    Both 'broker_statement' and 'preclearance' must be registered first via register_compliance_file.
    Returns a structured JSON summary for narration plus attaches a downloadable Excel report.
    """
    session = get_sandbox_session()
    if session is None:
        return json.dumps({"error": "Sandbox not available."})

    registry = await _read_registry(session)
    missing = [r for r in ("broker_statement", "preclearance") if r not in registry]
    if missing:
        return json.dumps({"error": f"Missing required files: {', '.join(missing)}. Please upload them first."})

    script_source = (Path(__file__).parent.parent / "compliance_check.py").read_bytes()
    await _sbx_write(session, _SCRIPT_PATH, script_source)

    trades_path = registry["broker_statement"]["json_path"]
    pc_path = registry["preclearance"]["json_path"]
    cmd = f"python3 {_SCRIPT_PATH} {trades_path} {pc_path}"
    if "index_reference" in registry:
        cmd += f" {registry['index_reference']['json_path']}"

    stdout, stderr, code = await _sbx_run(session, cmd)
    if code != 0:
        return json.dumps({"error": f"Comparison script failed: {stderr}"})

    result = json.loads(stdout)

    report_bytes = await _generate_report(session, result)

    report_local = Path("/tmp/compliance_report.xlsx")
    report_local.write_bytes(report_bytes)
    file_element = cl.File(
        name="compliance_report.xlsx",
        path=str(report_local),
        display="inline",
    )
    await cl.Message(content="Compliance report ready:", elements=[file_element]).send()

    return json.dumps(result)
